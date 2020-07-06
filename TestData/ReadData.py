import openpyxl

workbook=openpyxl.load_workbook(r"C:\python\RobotFramework1\datasheet1.xlsx")

def fetch_number_of_row(sheetName):
    sheet=workbook[sheetName]
    return sheet.max_row

def fetch_cell_data(sheetName,row,cell):
    sheet=workbook[sheetName]
    cell=sheet.cell(int(row),int(cell))
    return cell.value


# print(fetch_number_of_row("TestData"))
# print(fetch_cell_data("TestData",1,2))

